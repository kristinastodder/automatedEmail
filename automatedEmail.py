#! python3
# Automated Project Email Reminder.py - Sends emails based on yes or no status in spreadsheet.

import openpyxl, smtplib, sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Open the spreadsheet and get the latest status.
wb = openpyxl.load_workbook('Email Program.xlsx')
sheet = wb['Sheet1']

# Check each person's status.
receiveEmail = {}
for r in range(2, sheet.max_row + 1):
    sendIt = sheet.cell(row=r, column=1).value
    if sendIt != 'no':
        name = sheet.cell(row=r, column=3).value
        email = sheet.cell(row=r, column=6).value
        receiveEmail[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
password = input('Enter your password:')
smtpObj.login('kristina.stodder@buildgc.com', password)


# Send out reminder emails.
for name, email in receiveEmail.items():
    from_addr = 'kristina.stodder@buildgc.com'
    to_addr = email
    cc_addr = 'estimatingpw@buildgc.com'
    
    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Cc'] = cc_addr
    msg['Subject'] = 'New Project at Westfield - SFC'
    body = "Hi %s,\n\n\
We have a new project called SFC - 144 LLW. Wenjuan invited you through Building Connected and I want to make sure that you received the invite.\n\n\
Please let me know if you are interested in bidding this one for us.\n\n\
It is due next Tuesday. \n\n\
Thank you,\n\n\
Kristina Stodder\n\
Estimating Coordinator\n\
BUILD GROUP\n\
Direct: 415-660-7519\n\
Estimating Fax: 415-366-2883" %(name)
    msg.attach(MIMEText(body, 'plain'))
    text = msg.as_string()
                    
    sendmailStatus = smtpObj.sendmail(from_addr, to_addr, text)

if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

# Cc estimatingpw@buildgc.com
for name, email in receiveEmail.items():
    from_addr = 'kristina.stodder@buildgc.com'
    to_addr = 'estimatingpw@buildgc.com'
    cc_addr = email
    
    msg = MIMEMultipart()
    msg['From'] = from_addr
    msg['To'] = cc_addr
    msg['Cc'] = to_addr
    msg['Subject'] = 'New Project at Westfield - SFC'
    body = "Hi %s,\n\n\
We have a new project called SFC - 144 LLW. Wenjuan invited you through Building Connected and I want to make sure that you received the invite.\n\n\
Please let me know if you are interested in bidding this one for us.\n\n\
It is due next Tuesday. \n\n\
Thank you,\n\n\
Kristina Stodder\n\
Estimating Coordinator\n\
BUILD GROUP\n\
Direct: 415-660-7519\n\
Estimating Fax: 415-366-2883" %(name)
    msg.attach(MIMEText(body, 'plain'))
    text = msg.as_string()
    
    print('Sending email to %s at %s and copying estimatingpw@buildgc.com' %(name, email))
                    
    sendmailStatus = smtpObj.sendmail(from_addr, to_addr, text)

if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()

print('Complete')

